import os
import io
import time
import zipfile
import re
import base64
import shutil
import requests
from datetime import datetime
import openpyxl
from PIL import Image
from docxtpl import DocxTemplate
import db

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(BASE_DIR, "TEMPLATE_placeholders.docx")
ASSET_DIR = os.path.join(BASE_DIR, "web_assets")
IS_VERCEL = bool(os.environ.get('VERCEL'))
OUTPUT_DIR = '/tmp/output' if IS_VERCEL else os.path.join(BASE_DIR, "output")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(ASSET_DIR, exist_ok=True)

def get_transparent_png_bytes() -> bytes:
    """Creates a 1x1 transparent PNG in memory for placeholder clearing"""
    img = Image.new("RGBA", (1, 1), (255, 255, 255, 0))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def calculate_initials(nama: str) -> str:
    """
    Calculates 2-letter initials from first word of name (e.g. 'ASEP JALAM' -> 'AS', 'GAYA NAYLA' -> 'GA')
    """
    clean = re.sub(r'[^A-Z\s]', '', nama.upper()).strip()
    words = clean.split()
    if words and len(words[0]) >= 2:
        return words[0][:2].upper()
    elif len(words) >= 2:
        return (words[0][0] + words[1][0]).upper()
    elif words and len(words[0]) == 1:
        return (words[0] + "X").upper()
    return "GA"

def generate_certificate_docx(data: dict, qr_bytes: bytes = None, output_filename: str = None) -> tuple:
    """
    Generates single official TELP Certificate DOCX using docxtpl and exact placeholder rules.
    """
    # Select template based on score report flag
    is_score_report = bool(data.get("is_score_report")) or data.get("kategori") == "score_report"
    score_report_template = os.path.join(BASE_DIR, "resources", "templates", "SCORE_REPORT_TEMPLATE.docx")

    if is_score_report and os.path.exists(score_report_template):
        target_template = score_report_template
    else:
        target_template = DEFAULT_TEMPLATE

    if not os.path.exists(target_template):
        raise FileNotFoundError(f"Template tidak ditemukan di: {target_template}")

    # Prepare QR Image
    temp_qr = None
    if qr_bytes and len(qr_bytes) > 50:
        temp_qr = os.path.join(OUTPUT_DIR, f"_temp_qr_{int(time.time()*1000)}.png")
        with open(temp_qr, "wb") as f:
            f.write(qr_bytes)
        qr_file = temp_qr
    else:
        qr_type = data.get("qr_type", "s1").lower()
        qr_file = os.path.join(ASSET_DIR, f"qr_{qr_type}.png")
        if not os.path.exists(qr_file):
            qr_file = os.path.join(ASSET_DIR, "qr_s1.png")
        if not os.path.exists(qr_file):
            qr_file = os.path.join(ASSET_DIR, "image5.jpeg")

    # Render docxtpl template
    doc = DocxTemplate(target_template)
    doc.render(data)

    # Replace QR Picture
    if os.path.exists(qr_file):
        try:
            doc.replace_pic("qr_code", qr_file)
        except Exception as e:
            pass

    # Save to in-memory buffer
    mem_buf = io.BytesIO()
    doc.save(mem_buf)
    mem_buf.seek(0)

    if temp_qr and os.path.exists(temp_qr):
        try:
            os.remove(temp_qr)
        except Exception:
            pass

    show_validation_logo = bool(data.get("show_validation_logo", True))
    show_cap_ttd = bool(data.get("show_cap_ttd", True))

    transparent_bytes = get_transparent_png_bytes()
    out_zip_buf = io.BytesIO()

    with zipfile.ZipFile(mem_buf, "r") as zin:
        with zipfile.ZipFile(out_zip_buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == "word/media/image7.png":
                    if not show_validation_logo:
                        content = transparent_bytes
                elif item.filename == "word/media/image6.png":
                    if not show_cap_ttd:
                        content = transparent_bytes
                zout.writestr(item, content)

    out_docx_bytes = out_zip_buf.getvalue()

    if not output_filename:
        safe_no = str(data.get("no_sertifikat", "sertifikat")).replace("/", "_").replace("\\", "_")
        output_filename = f"SERTIFIKAT_{safe_no}.docx"

    if os.path.isabs(output_filename):
        docx_path = output_filename
    elif output_filename.startswith(OUTPUT_DIR) or output_filename.startswith("output"):
        docx_path = os.path.abspath(output_filename)
    else:
        docx_path = os.path.join(OUTPUT_DIR, output_filename)

    os.makedirs(os.path.dirname(docx_path), exist_ok=True)
    with open(docx_path, "wb") as f:
        f.write(out_docx_bytes)

    return docx_path, out_docx_bytes

def convert_docx_to_pdf(docx_path: str, pdf_path: str = None) -> str:
    """
    Converts DOCX to PDF using MS Word COM automation on Windows.
    """
    if not pdf_path:
        pdf_path = docx_path.rsplit('.', 1)[0] + '.pdf'

    abs_docx = os.path.abspath(docx_path)
    abs_pdf = os.path.abspath(pdf_path)

    if not os.path.exists(abs_docx):
        print(f"Error: DOCX file not found: {abs_docx}")
        return None

    if os.path.exists(abs_pdf):
        try:
            os.remove(abs_pdf)
        except Exception:
            pass

    try:
        import pythoncom
        pythoncom.CoInitialize()
    except (ImportError, Exception):
        # pythoncom / Word COM only exists on Windows with Microsoft Word.
        # Safely skip on Linux / Vercel Serverless.
        return None

    word = None
    doc = None
    try:
        import win32com.client
        word = win32com.client.Dispatch("Word.Application")
        try:
            word.Visible = False
        except Exception:
            pass
        
        doc = word.Documents.Open(
            FileName=abs_docx,
            ConfirmConversions=False,
            ReadOnly=True,
            AddToRecentFiles=False
        )
        doc.SaveAs(FileName=abs_pdf, FileFormat=17) # 17 = wdFormatPDF
        doc.Close(SaveChanges=0)
        doc = None
        word.Quit()
        word = None

        if os.path.exists(abs_pdf) and os.path.getsize(abs_pdf) > 0:
            return abs_pdf
    except Exception as e:
        print(f"Direct Word COM error on {os.path.basename(docx_path)}: {e}")
    finally:
        try:
            if doc:
                doc.Close(0)
        except Exception:
            pass
        try:
            if word:
                word.Quit()
        except Exception:
            pass

    # Fallback docx2pdf
    try:
        from docx2pdf import convert
        convert(abs_docx, abs_pdf)
        if os.path.exists(abs_pdf) and os.path.getsize(abs_pdf) > 0:
            return abs_pdf
    except Exception as e2:
        print(f"docx2pdf fallback error: {e2}")

    return None

def extract_gdrive_folder_id(val: str) -> str:
    """Extract clean folder ID from raw ID or full Google Drive URL"""
    if not val:
        return ""
    val = str(val).strip()
    if "/folders/" in val:
        match = re.search(r'/folders/([a-zA-Z0-9_-]+)', val)
        if match:
            return match.group(1)
    if "id=" in val:
        match = re.search(r'id=([a-zA-Z0-9_-]+)', val)
        if match:
            return match.group(1)
    return val

def upload_to_gdrive(file_path: str, gdrive_url: str, folder_id: str, custom_filename: str = None) -> dict:
    """
    Uploads a file (PDF or DOCX) to Google Drive via Google Apps Script Web App.
    Fully compatible with user's existing script (file_base64 & status: success).
    """
    if not file_path or not os.path.exists(file_path):
        return {'success': False, 'error': f'File tidak ditemukan: {file_path}'}
    
    clean_url = str(gdrive_url or '').strip()
    clean_folder = extract_gdrive_folder_id(folder_id)

    if not clean_url:
        return {'success': False, 'error': 'URL Web App Google Drive belum diisi'}
    if not clean_folder:
        return {'success': False, 'error': 'Folder ID Google Drive belum diisi untuk kategori ini'}

    try:
        import requests
        with open(file_path, 'rb') as f:
            b64_data = base64.b64encode(f.read()).decode('utf-8')

        filename = custom_filename or os.path.basename(file_path)
        mime_type = 'application/pdf' if filename.lower().endswith('.pdf') else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'

        # Send BOTH file_base64 (user script) and file_data (universal)
        payload = {
            'action': 'upload',
            'filename': filename,
            'file_base64': b64_data,
            'file_data': b64_data,
            'folder_id': clean_folder,
            'mime_type': mime_type
        }
        res = requests.post(clean_url, json=payload, timeout=45)
        try:
            res_json = res.json()
            is_success = res_json.get('status') == 'success' or res_json.get('success') is True
            return {
                'success': is_success,
                'file_id': res_json.get('file_id'),
                'url': res_json.get('url') or res_json.get('file_url'),
                'error': res_json.get('message') or res_json.get('error') if not is_success else None,
                'raw': res_json
            }
        except Exception:
            return {'success': res.status_code == 200, 'status': res.status_code, 'text': res.text[:200]}
    except Exception as e:
        return {'success': False, 'error': str(e)}

def generate_certificate_full(
    data: dict,
    generate_pdf: bool = True,
    generate_polos: bool = True,
    gdrive_settings: dict = None
) -> dict:
    """
    Generates all versions of a certificate:
    1. DOCX Resmi (with cap & validation)
    2. PDF Resmi (with cap & validation)
    3. DOCX Polos (without cap & validation)
    4. PDF Polos (without cap & validation)
    5. Optionally auto-uploads PDF Resmi to Google Drive based on qr_type (s1, pasca, umum).
    """
    safe_no = str(data.get("no_sertifikat", "sertifikat")).replace("/", "_").replace("\\", "_")
    safe_nama = re.sub(r'\s+', '_', str(data.get("nama", "peserta")).strip().upper())[:20]

    # 1. DOCX & PDF Resmi (Digital)
    resmi_data = dict(data)
    resmi_data['show_validation_logo'] = bool(data.get('show_validation_logo', data.get('show_validation', True)))
    resmi_data['show_cap_ttd'] = bool(data.get('show_cap_ttd', data.get('show_cap', True)))

    docx_resmi_filename = f"SERTIFIKAT_{safe_no}_{safe_nama}.docx"
    docx_resmi_path, _ = generate_certificate_docx(resmi_data, output_filename=docx_resmi_filename)

    pdf_resmi_filename = f"SERTIFIKAT_{safe_no}_{safe_nama}.pdf"
    pdf_resmi_path = os.path.join(OUTPUT_DIR, pdf_resmi_filename)
    if generate_pdf:
        convert_docx_to_pdf(docx_resmi_path, pdf_resmi_path)

    # 2. DOCX & PDF Polos (Untuk Cetak Blanko)
    docx_polos_filename = f"SERTIFIKAT_{safe_no}_{safe_nama}_CETAK_POLOS.docx"
    docx_polos_path = os.path.join(OUTPUT_DIR, docx_polos_filename)
    pdf_polos_filename = f"SERTIFIKAT_{safe_no}_{safe_nama}_CETAK_POLOS.pdf"
    pdf_polos_path = os.path.join(OUTPUT_DIR, pdf_polos_filename)

    if generate_polos:
        polos_data = dict(data)
        polos_data['show_validation_logo'] = False
        polos_data['show_cap_ttd'] = False
        generate_certificate_docx(polos_data, output_filename=docx_polos_filename)
        if generate_pdf:
            convert_docx_to_pdf(docx_polos_path, pdf_polos_path)

    # 3. Auto Upload PDF Resmi to Google Drive if active
    gdrive_upload = None
    if not gdrive_settings or not gdrive_settings.get('url'):
        db_s = db.get_all_app_settings()
        if db_s.get('gdrive_url'):
            gdrive_settings = {
                'url': db_s.get('gdrive_url', ''),
                'folder_s1': db_s.get('gdrive_folder_s1', ''),
                'folder_pasca': db_s.get('gdrive_folder_pasca', ''),
                'folder_umum': db_s.get('gdrive_folder_umum', ''),
                'auto_sync': db_s.get('gdrive_auto_sync') == '1'
            }

    if gdrive_settings and gdrive_settings.get('auto_sync'):
        qr_type = str(data.get('qr_type', 's1')).lower()
        folder_id = gdrive_settings.get(f'folder_{qr_type}') or gdrive_settings.get('folder_id')
        gdrive_url = gdrive_settings.get('url') or gdrive_settings.get('gdrive_url')
        
        if not gdrive_url:
            gdrive_upload = {'success': False, 'error': 'URL Web App Google Drive belum diatur di menu Integrasi.'}
        elif not folder_id:
            gdrive_upload = {'success': False, 'error': f'ID Folder Google Drive untuk kategori {qr_type.upper()} belum diisi di menu Integrasi Google Drive.'}
        elif pdf_resmi_path and os.path.exists(pdf_resmi_path):
            safe_nama_gd = re.sub(r'[\/\\:*?"<>|]', '', str(data.get('nama', 'PESERTA'))).strip().upper()
            safe_cert_no_gd = str(data.get('no_sertifikat', 'CERT')).replace('/', '_').replace('\\', '_').strip()
            kode_awal_gd = re.split(r'[A-Za-z]', safe_cert_no_gd)[0] if re.search(r'[A-Za-z]', safe_cert_no_gd) else safe_cert_no_gd
            if not kode_awal_gd:
                kode_awal_gd = safe_cert_no_gd
            gdrive_filename = f"{safe_nama_gd}-{kode_awal_gd}.pdf"
            
            gdrive_upload = upload_to_gdrive(pdf_resmi_path, gdrive_url, folder_id, custom_filename=gdrive_filename)
        else:
            gdrive_upload = {'success': False, 'error': 'PDF Resmi tidak ditemukan untuk diunggah.'}

    # Copy to configured custom directory with structured folders:
    # output_dir / QR_TYPE / TANGGAL / [NAMA-NO_SERTIFIKAT] / [NAMA-KODE_AWAL].docx + [NAMA-KODE_AWAL].pdf
    saved_to_dirs = {}
    try:
        app_settings = db.get_all_app_settings()
        output_dir = app_settings.get('output_dir') or ''

        if output_dir:
            qr_folder = qr_type.upper() if qr_type else 'S1'
            date_folder = datetime.now().strftime('%Y-%m-%d')
            safe_nama = re.sub(r'[\/\\:*?"<>|]', '', str(data.get('nama', 'PESERTA'))).strip().upper()
            safe_cert_no = str(data.get('no_sertifikat', 'CERT')).replace('/', '_').replace('\\', '_').strip()
            
            # Kode awal sebelum kode lembaga (misal: 135512 dari 135512SULCUN58038GA)
            kode_awal = re.split(r'[A-Za-z]', safe_cert_no)[0] if re.search(r'[A-Za-z]', safe_cert_no) else safe_cert_no
            if not kode_awal:
                kode_awal = safe_cert_no

            folder_peserta = f"{safe_nama}-{safe_cert_no}"
            cert_folder = os.path.join(output_dir, qr_folder, date_folder, folder_peserta)
            os.makedirs(cert_folder, exist_ok=True)

            target_word_name = f"{safe_nama}-{kode_awal}.docx"
            target_pdf_name = f"{safe_nama}-{kode_awal}.pdf"

            if os.path.exists(docx_resmi_path):
                dst_word = os.path.join(cert_folder, target_word_name)
                shutil.copy2(docx_resmi_path, dst_word)
                saved_to_dirs['word'] = dst_word

            if pdf_polos_path and os.path.exists(pdf_polos_path):
                dst_pdf = os.path.join(cert_folder, target_pdf_name)
                shutil.copy2(pdf_polos_path, dst_pdf)
                saved_to_dirs['pdf_polos'] = dst_pdf

            saved_to_dirs['folder'] = cert_folder
    except Exception as e:
        print(f"Error copying to custom directories: {e}")

    # Save to SQLite DB
    db_record = dict(resmi_data)
    db_record['docx_path'] = docx_resmi_path
    if pdf_resmi_path and os.path.exists(pdf_resmi_path):
        db_record['pdf_path'] = pdf_resmi_path
    db.save_certificate_db(db_record)

    return {
        'success': True,
        'no_sertifikat': data.get('no_sertifikat'),
        'nama': data.get('nama'),
        'docx_resmi': {
            'filename': docx_resmi_filename,
            'url': f"/api/v1/certificates/download/{docx_resmi_filename}"
        },
        'pdf_resmi': {
            'filename': pdf_resmi_filename,
            'url': f"/api/v1/certificates/download/{pdf_resmi_filename}" if os.path.exists(pdf_resmi_path) else None
        },
        'docx_polos': {
            'filename': docx_polos_filename,
            'url': f"/api/v1/certificates/download/{docx_polos_filename}" if os.path.exists(docx_polos_path) else None
        },
        'pdf_polos': {
            'filename': pdf_polos_filename,
            'url': f"/api/v1/certificates/download/{pdf_polos_filename}" if os.path.exists(pdf_polos_path) else None
        },
        'saved_to_dirs': saved_to_dirs,
        'gdrive_upload': gdrive_upload
    }

def generate_pdf_polos_stream(data: dict) -> str:
    """Generates or retrieves a plain PDF (without cap and without validation) for on-demand printing."""
    safe_no = str(data.get("no_sertifikat", "sertifikat")).replace("/", "_").replace("\\", "_")
    safe_nama = re.sub(r'\s+', '_', str(data.get("nama", "peserta")).strip().upper())[:20]
    
    pdf_filename = f"SERTIFIKAT_{safe_no}_{safe_nama}_CETAK_POLOS.pdf"
    pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)
    
    if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
        return pdf_path

    polos_data = dict(data)
    polos_data['show_validation_logo'] = False
    polos_data['show_cap_ttd'] = False
    
    docx_filename = f"SERTIFIKAT_{safe_no}_{safe_nama}_CETAK_POLOS.docx"
    docx_path = os.path.join(OUTPUT_DIR, docx_filename)
    if not os.path.exists(docx_path):
        docx_path, _ = generate_certificate_docx(polos_data, output_filename=docx_filename)
    
    convert_docx_to_pdf(docx_path, pdf_path)
    return pdf_path

# =============================================================
# BULK EXCEL STUDIO GENERATOR ENGINE
# =============================================================
MONTH_MAP = {
    'januari': '01', 'january': '01', 'jan': '01',
    'februari': '02', 'february': '02', 'feb': '02',
    'maret': '03', 'march': '03', 'mar': '03',
    'april': '04', 'apr': '04',
    'mei': '05', 'may': '05',
    'juni': '06', 'june': '06', 'jun': '06',
    'juli': '07', 'july': '07', 'jul': '07',
    'agustus': '08', 'august': '08', 'agu': '08', 'agt': '08', 'aug': '08',
    'september': '09', 'sep': '09', 'sept': '09',
    'oktober': '10', 'october': '10', 'okt': '10', 'oct': '10',
    'november': '11', 'nov': '11',
    'desember': '12', 'december': '12', 'des': '12', 'dec': '12'
}

def extract_and_format_dob(val, default="04/01/2009"):
    if val is None or val == "":
        return default
    if hasattr(val, 'strftime'):
        return val.strftime("%d/%m/%Y")
    
    val_str = str(val).strip()
    if not val_str or val_str.lower() == 'none':
        return default

    # If it contains comma or slash for TTL (e.g. "TASIKMALAYA, 04/01/2009" or "BANDUNG, 15 JANUARI 2002")
    if ',' in val_str:
        parts = val_str.split(',', 1)
        val_str = parts[1].strip()

    # 1. Check for text month (e.g. "15 Januari 2002" or "5 Mei 2003" or "25-Agustus-2001")
    text_month_match = re.search(r'(\d{1,2})[\s\-_]+([a-zA-Z]+)[\s\-_]+(\d{4})', val_str)
    if text_month_match:
        day = int(text_month_match.group(1))
        month_name = text_month_match.group(2).lower()
        year = text_month_match.group(3)
        month_num = MONTH_MAP.get(month_name, '01')
        return f"{day:02d}/{month_num}/{year}"

    # 2. Check for DD/MM/YYYY or DD-MM-YYYY
    dmy_match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', val_str)
    if dmy_match:
        day = int(dmy_match.group(1))
        month = int(dmy_match.group(2))
        year = dmy_match.group(3)
        return f"{day:02d}/{month:02d}/{year}"

    # 3. Check for YYYY-MM-DD or YYYY/MM/DD
    ymd_match = re.search(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', val_str)
    if ymd_match:
        year = ymd_match.group(1)
        month = int(ymd_match.group(2))
        day = int(ymd_match.group(3))
        return f"{day:02d}/{month:02d}/{year}"

    # Fallback if 8 digits like 20090104 or 04012009
    if re.match(r'^\d{8}$', val_str):
        if int(val_str[:4]) > 1900: # YYYYMMDD
            return f"{val_str[6:8]}/{val_str[4:6]}/{val_str[:4]}"
        else: # DDMMYYYY
            return f"{val_str[:2]}/{val_str[2:4]}/{val_str[4:]}"

def parse_excel_columns(rows):
    header_idx = 0
    col_map = {}
    best_match_count = 0

    # Scan first 15 rows to find the row with the most matching certificate header columns
    for idx in range(min(15, len(rows))):
        row = rows[idx]
        if not row: continue
        current_map = {}
        for col_i, cell in enumerate(row):
            if cell is None: continue
            c_str = re.sub(r'[^a-zA-Z0-9/&_\s]', '', str(cell).lower()).strip()
            
            # 1. Nama Peserta
            if any(k == c_str or k in c_str for k in ["nama", "name", "mahasiswa", "peserta", "nama lengkap", "student"]) and 'nama' not in current_map:
                current_map['nama'] = col_i
            # 2. Tanggal Lahir / TTL
            elif any(k == c_str or k in c_str for k in ["ttl", "lahir", "birth", "dob", "tgl_lhr", "tgl. lahir", "tgl lahir", "tempat tanggal lahir", "tempat, tanggal lahir", "tempat / tgl", "tempat/tgl", "t.t.l", "tanggal lahir"]) and 'tgl_lahir' not in current_map:
                current_map['tgl_lahir'] = col_i
            # 3. Jenis Kelamin / Gender
            elif any(k == c_str or k in c_str for k in ["gender", "sex", "kelamin", "jk", "l/p", "jenis kelamin"]) and 'gender' not in current_map:
                current_map['gender'] = col_i
            # 4. Tanggal Tes (Explicit)
            elif any(k == c_str or k in c_str for k in ["tgl_tes", "tgl tes", "tgl ujian", "tanggal tes", "tanggal ujian", "test date", "exam date", "tgl_ujian"]) and 'tgl_tes' not in current_map:
                current_map['tgl_tes'] = col_i
            # 5. Listening (LC, Section 1, Listening)
            elif any(k == c_str or k in c_str for k in ["listening", "score_l", "sec1", "section 1", "section_1", "lis", "lc", "l/c", "l.c", "listening comprehension", "sec 1"]) and 'listening' not in current_map:
                current_map['listening'] = col_i
            # 6. Structure (SW, Section 2, Structure)
            elif any(k == c_str or k in c_str for k in ["structure", "score_s", "sec2", "section 2", "section_2", "str", "sw", "s/w", "s.w", "structure and written expression", "structure & written", "sec 2"]) and 'structure' not in current_map:
                current_map['structure'] = col_i
            # 7. Reading (RC, Section 3, Reading)
            elif any(k == c_str or k in c_str for k in ["reading", "score_r", "sec3", "section 3", "section_3", "read", "rc", "r/c", "r.c", "reading comprehension", "sec 3"]) and 'reading' not in current_map:
                current_map['reading'] = col_i
            # 8. Overall / Total
            elif any(k == c_str or k in c_str for k in ["overall", "total", "score", "skor", "nilai", "overall score", "band score", "total score", "skor total", "nilai total", "total band", "tot"]) and 'overall' not in current_map:
                current_map['overall'] = col_i

        if len(current_map) > best_match_count:
            best_match_count = len(current_map)
            header_idx = idx
            col_map = current_map

    # If tgl_lahir was still not found by header keyword, auto-detect from data rows
    if 'tgl_lahir' not in col_map and len(rows) > header_idx + 1:
        for c_i, val in enumerate(rows[header_idx + 1]):
            if c_i not in col_map.values() and val is not None:
                val_s = str(val).lower()
                if hasattr(val, 'strftime') or any(m in val_s for m in ['19', '20', 'jan', 'feb', 'mar', 'apr', 'mei', 'jun', 'jul', 'agu', 'sep', 'okt', 'nov', 'des', '/', '-']) and len(val_s) >= 4:
                    col_map['tgl_lahir'] = c_i
                    break

    return header_idx, col_map

# =============================================================
# BULK EXCEL STUDIO GENERATOR ENGINE
# =============================================================
def generate_bulk_certificates_from_excel(
    excel_path: str,
    kd_doc: str = "135512",
    kd_lem: str = "SULCUN58",
    start_urut: int = 1,
    qr_type: str = "s1",
    show_validation: bool = True,
    show_cap: bool = True,
    default_tgl_tes: str = "19/12/2025"
) -> dict:
    """
    Parses raw student Excel file and generates bulk certificates DOCX,
    saving to DB and zipping into a single download package.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel kosong")

    header_idx, col_map = parse_excel_columns(rows)
    if 'nama' not in col_map:
        col_map = {'nama': 0, 'tgl_lahir': 1, 'gender': 2, 'listening': 3, 'structure': 4, 'reading': 5, 'overall': 6}

    batch_id = f"batch_{int(time.time())}"
    batch_dir = os.path.join(OUTPUT_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    generated_records = []
    current_urut = start_urut

    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or not any(row): continue
        
        raw_nama = str(row[col_map.get('nama', 0)] or '').strip()
        if not raw_nama or raw_nama.lower() == 'none': continue

        clean_nama = re.sub(r'\s+', ' ', raw_nama).upper()
        inisial = calculate_initials(clean_nama)
        str_urut = f"{current_urut:03d}"
        full_no = f"{kd_doc}{kd_lem}{str_urut}{inisial}"

        # Tanggal Lahir (Formatted strictly DD/MM/YYYY)
        if 'tgl_lahir' in col_map and col_map['tgl_lahir'] < len(row):
            tgl_lahir = extract_and_format_dob(row[col_map['tgl_lahir']], '04/01/2009')
        else:
            tgl_lahir = '04/01/2009'
        
        # Tanggal Tes: only if column explicitly found, else use default_tgl_tes
        if 'tgl_tes' in col_map and col_map['tgl_tes'] < len(row):
            tgl_tes = extract_and_format_dob(row[col_map['tgl_tes']], default_tgl_tes)
        else:
            tgl_tes = default_tgl_tes

        # Calculate expiry (+2 years)
        parts = tgl_tes.split('/')
        if len(parts) == 3:
            try:
                yyyy = int(parts[2]) + 2
                valid_until = f"{parts[0]}/{parts[1]}/{yyyy}"
            except Exception:
                valid_until = "19/12/2027"
        else:
            valid_until = "19/12/2027"

        # Gender
        raw_gender = ''
        if 'gender' in col_map and col_map['gender'] < len(row):
            raw_gender = str(row[col_map['gender']] or 'F').strip().upper()
        gender = 'M (Male)' if raw_gender.startswith('M') or raw_gender.startswith('L') else 'F (Female)'

        # Scores
        listening = str(row[col_map['listening']] if 'listening' in col_map and col_map['listening'] < len(row) and row[col_map['listening']] is not None else 56).split('.')[0]
        structure = str(row[col_map['structure']] if 'structure' in col_map and col_map['structure'] < len(row) and row[col_map['structure']] is not None else 53).split('.')[0]
        reading = str(row[col_map['reading']] if 'reading' in col_map and col_map['reading'] < len(row) and row[col_map['reading']] is not None else 50).split('.')[0]
        overall = str(row[col_map['overall']] if 'overall' in col_map and col_map['overall'] < len(row) and row[col_map['overall']] is not None else 530).split('.')[0]

        cert_data = {
            'no_sertifikat': full_no,
            'nama': clean_nama,
            'tanggal_lahir': tgl_lahir,
            'jenis_kelamin': gender,
            'negara': 'Indonesia',
            'bahasa': 'Indonesian',
            'tanggal_tes': tgl_tes,
            'berlaku_sampai': valid_until,
            'listening': listening,
            'structure': structure,
            'reading': reading,
            'overall': overall,
            'qr_type': qr_type,
            'show_validation_logo': show_validation,
            'show_cap_ttd': show_cap
        }

        # Generate DOCX
        safe_filename = f"SERTIFIKAT_{full_no}_{clean_nama.replace(' ', '_')[:20]}.docx"
        out_single_path = os.path.join(batch_dir, safe_filename)
        generate_certificate_docx(cert_data, output_filename=out_single_path)

        cert_data['docx_path'] = out_single_path
        db.save_certificate_db(cert_data)

        generated_records.append({
            'no_sertifikat': full_no,
            'nama': clean_nama,
            'overall': overall,
            'filename': safe_filename
        })
        current_urut += 1

    # Create ZIP archive containing all DOCX files
    zip_filename = f"PAKET_BULK_SERTIFIKAT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_filename)

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(batch_dir):
            for file in files:
                zipf.write(os.path.join(root, file), arcname=file)

    return {
        'success': True,
        'count': len(generated_records),
        'zip_filename': zip_filename,
        'zip_url': f"/api/v1/certificates/download/{zip_filename}",
        'records': generated_records
    }

def parse_bulk_excel(
    excel_path: str,
    kd_doc: str = "135512",
    kd_lem: str = "SULCUN58",
    start_urut: int = 1,
    qr_type: str = "s1",
    default_tgl_tes: str = "19/12/2025",
    custom_col_map: dict = None,
    custom_header_idx: int = None
) -> dict:
    """
    Parses Excel file and returns structured records list + headers for interactive studio review/editing.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel kosong")

    detected_header_idx, detected_col_map = parse_excel_columns(rows)
    
    header_idx = custom_header_idx if custom_header_idx is not None else detected_header_idx
    col_map = custom_col_map if custom_col_map is not None else detected_col_map

    if 'nama' not in col_map:
        col_map['nama'] = 0

    parsed_records = []
    current_urut = start_urut

    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or not any(row): continue
        raw_nama = str(row[col_map.get('nama', 0)] or '').strip()
        if not raw_nama or raw_nama.lower() == 'none': continue

        clean_nama = re.sub(r'\s+', ' ', raw_nama).upper()
        inisial = calculate_initials(clean_nama)
        str_urut = f"{current_urut:03d}"
        full_no = f"{kd_doc}{kd_lem}{str_urut}{inisial}"

        # Tanggal Lahir / TTL (Formatted strictly DD/MM/YYYY)
        if 'tgl_lahir' in col_map and col_map['tgl_lahir'] is not None and col_map['tgl_lahir'] < len(row):
            tgl_lahir = extract_and_format_dob(row[col_map['tgl_lahir']], '04/01/2009')
        else:
            tgl_lahir = '04/01/2009'

        # Tanggal Tes: only if column explicitly found, else use default_tgl_tes
        if 'tgl_tes' in col_map and col_map['tgl_tes'] is not None and col_map['tgl_tes'] < len(row):
            tgl_tes = extract_and_format_dob(row[col_map['tgl_tes']], default_tgl_tes)
        else:
            tgl_tes = default_tgl_tes

        parts = tgl_tes.split('/')
        if len(parts) == 3:
            try:
                yyyy = int(parts[2]) + 2
                valid_until = f"{parts[0]}/{parts[1]}/{yyyy}"
            except Exception:
                valid_until = "19/12/2027"
        else:
            valid_until = "19/12/2027"

        raw_gender = ''
        if 'gender' in col_map and col_map['gender'] is not None and col_map['gender'] < len(row):
            raw_gender = str(row[col_map['gender']] or 'F').strip().upper()
        gender = 'M (Male)' if raw_gender.startswith('M') or raw_gender.startswith('L') else 'F (Female)'

        listening = str(row[col_map['listening']] if 'listening' in col_map and col_map['listening'] is not None and col_map['listening'] < len(row) and row[col_map['listening']] is not None else 56).split('.')[0]
        structure = str(row[col_map['structure']] if 'structure' in col_map and col_map['structure'] is not None and col_map['structure'] < len(row) and row[col_map['structure']] is not None else 53).split('.')[0]
        reading = str(row[col_map['reading']] if 'reading' in col_map and col_map['reading'] is not None and col_map['reading'] < len(row) and row[col_map['reading']] is not None else 50).split('.')[0]
        overall = str(row[col_map['overall']] if 'overall' in col_map and col_map['overall'] is not None and col_map['overall'] < len(row) and row[col_map['overall']] is not None else 530).split('.')[0]

        parsed_records.append({
            'index': len(parsed_records),
            'no_urut': current_urut,
            'kd_doc': kd_doc,
            'kd_lem': kd_lem,
            'inisial': inisial,
            'no_sertifikat': full_no,
            'nama': clean_nama,
            'tanggal_lahir': tgl_lahir,
            'jenis_kelamin': gender,
            'negara': 'Indonesia',
            'bahasa': 'Indonesian',
            'tanggal_tes': tgl_tes,
            'berlaku_sampai': valid_until,
            'listening': listening,
            'structure': structure,
            'reading': reading,
            'overall': overall,
            'qr_type': qr_type,
            'show_validation': True,
            'show_cap': True
        })
        current_urut += 1

    headers = [str(c or f"Kolom {i+1}").strip() for i, c in enumerate(rows[header_idx])]

    return {
        'records': parsed_records,
        'headers': headers,
        'col_map': col_map,
        'header_idx': header_idx,
        'total_rows': len(parsed_records)
    }

def generate_bulk_from_records(
    records: list,
    generate_pdf: bool = True,
    generate_polos: bool = True,
    gdrive_settings: dict = None
) -> dict:
    """
    Generates all certificates (PDF & DOCX for both Official and Polos formats),
    zips them into organized directories, and auto-uploads Official PDFs to Google Drive.
    """
    if not records:
        raise ValueError("Daftar peserta kosong")

    batch_id = f"batch_{int(time.time())}"
    batch_dir = os.path.join(OUTPUT_DIR, batch_id)
    dir_pdf_resmi = os.path.join(batch_dir, "1_PDF_RESMI_DIGITAL")
    dir_pdf_polos = os.path.join(batch_dir, "2_PDF_CETAK_POLOS")
    dir_docx_resmi = os.path.join(batch_dir, "3_DOCX_RESMI")
    dir_docx_polos = os.path.join(batch_dir, "4_DOCX_CETAK_POLOS")

    os.makedirs(dir_pdf_resmi, exist_ok=True)
    os.makedirs(dir_pdf_polos, exist_ok=True)
    os.makedirs(dir_docx_resmi, exist_ok=True)
    os.makedirs(dir_docx_polos, exist_ok=True)

    generated_records = []
    gdrive_uploaded_count = 0

    if not gdrive_settings or not gdrive_settings.get('url'):
        db_s = db.get_all_app_settings()
        if db_s.get('gdrive_url'):
            gdrive_settings = {
                'url': db_s.get('gdrive_url', ''),
                'folder_s1': db_s.get('gdrive_folder_s1', ''),
                'folder_pasca': db_s.get('gdrive_folder_pasca', ''),
                'folder_umum': db_s.get('gdrive_folder_umum', ''),
                'auto_sync': db_s.get('gdrive_auto_sync') == '1'
            }

    for item in records:
        nama = str(item.get('nama', '')).upper()
        safe_nama = re.sub(r'\s+', '_', nama)[:20]
        full_no = str(item.get('no_sertifikat', ''))
        safe_no = str(full_no).replace('/', '_').replace('\\', '_')
        qr_type = str(item.get('qr_type', 's1')).lower()

        cert_data_resmi = {
            'no_sertifikat': full_no,
            'nama': nama,
            'tanggal_lahir': item.get('tanggal_lahir', '04/01/2009'),
            'jenis_kelamin': item.get('jenis_kelamin', 'F (Female)'),
            'negara': item.get('negara', 'Indonesia'),
            'bahasa': item.get('bahasa', 'Indonesian'),
            'tanggal_tes': item.get('tanggal_tes', '19/12/2025'),
            'berlaku_sampai': item.get('berlaku_sampai', '19/12/2027'),
            'listening': str(item.get('listening', 56)),
            'structure': str(item.get('structure', 53)),
            'reading': str(item.get('reading', 50)),
            'overall': str(item.get('overall', 530)),
            'qr_type': qr_type,
            'show_validation_logo': bool(item.get('show_validation', True)),
            'show_cap_ttd': bool(item.get('show_cap', True))
        }

        # 1. Generate DOCX & PDF Resmi
        docx_resmi_path = os.path.join(dir_docx_resmi, f"SERTIFIKAT_{safe_no}_{safe_nama}.docx")
        generate_certificate_docx(cert_data_resmi, output_filename=docx_resmi_path)

        pdf_resmi_path = os.path.join(dir_pdf_resmi, f"SERTIFIKAT_{safe_no}_{safe_nama}.pdf")
        if generate_pdf:
            convert_docx_to_pdf(docx_resmi_path, pdf_resmi_path)

        # 2. Generate DOCX & PDF Polos (Untuk Cetak Blanko)
        if generate_polos:
            cert_data_polos = dict(cert_data_resmi)
            cert_data_polos['show_validation_logo'] = False
            cert_data_polos['show_cap_ttd'] = False

            docx_polos_path = os.path.join(dir_docx_polos, f"SERTIFIKAT_{safe_no}_{safe_nama}_CETAK_POLOS.docx")
            generate_certificate_docx(cert_data_polos, output_filename=docx_polos_path)

            if generate_pdf:
                pdf_polos_path = os.path.join(dir_pdf_polos, f"SERTIFIKAT_{safe_no}_{safe_nama}_CETAK_POLOS.pdf")
                convert_docx_to_pdf(docx_polos_path, pdf_polos_path)

        # 3. Google Drive Auto-Sync (Folder based on qr_type: S1, Pasca, Umum)
        if gdrive_settings and gdrive_settings.get('auto_sync') and os.path.exists(pdf_resmi_path):
            folder_id = gdrive_settings.get(f'folder_{qr_type}') or gdrive_settings.get('folder_id')
            gdrive_url = gdrive_settings.get('url') or gdrive_settings.get('gdrive_url')
            if gdrive_url and folder_id:
                safe_nama_bulk = re.sub(r'[\/\\:*?"<>|]', '', nama).strip().upper()
                safe_cert_no_bulk = str(full_no).replace('/', '_').replace('\\', '_').strip()
                kode_awal_bulk = re.split(r'[A-Za-z]', safe_cert_no_bulk)[0] if re.search(r'[A-Za-z]', safe_cert_no_bulk) else safe_cert_no_bulk
                if not kode_awal_bulk:
                    kode_awal_bulk = safe_cert_no_bulk
                gdrive_filename_bulk = f"{safe_nama_bulk}-{kode_awal_bulk}.pdf"

                upload_res = upload_to_gdrive(pdf_resmi_path, gdrive_url, folder_id, custom_filename=gdrive_filename_bulk)
                if upload_res.get('success'):
                    gdrive_uploaded_count += 1
                else:
                    gdrive_failed_count += 1
            else:
                gdrive_skipped_count += 1

        # 4. Copy to configured custom directory with structured folders:
        # output_dir / QR_TYPE / TANGGAL / [NAMA-NO_SERTIFIKAT] / [NAMA-KODE_AWAL].docx + [NAMA-KODE_AWAL].pdf
        try:
            app_settings = db.get_all_app_settings()
            custom_output_dir = app_settings.get('output_dir') or ''
            if custom_output_dir:
                qr_folder = qr_type.upper() if qr_type else 'S1'
                date_folder = datetime.now().strftime('%Y-%m-%d')
                safe_nama_clean = re.sub(r'[\/\\:*?"<>|]', '', nama).strip().upper()
                safe_cert_no_clean = str(full_no).replace('/', '_').replace('\\', '_').strip()
                kode_awal = re.split(r'[A-Za-z]', safe_cert_no_clean)[0] if re.search(r'[A-Za-z]', safe_cert_no_clean) else safe_cert_no_clean
                if not kode_awal:
                    kode_awal = safe_cert_no_clean

                folder_peserta = f"{safe_nama_clean}-{safe_cert_no_clean}"
                cert_folder = os.path.join(custom_output_dir, qr_folder, date_folder, folder_peserta)
                os.makedirs(cert_folder, exist_ok=True)

                target_word_name = f"{safe_nama_clean}-{kode_awal}.docx"
                target_pdf_name = f"{safe_nama_clean}-{kode_awal}.pdf"

                if os.path.exists(docx_resmi_path):
                    shutil.copy2(docx_resmi_path, os.path.join(cert_folder, target_word_name))
                if generate_polos and os.path.exists(pdf_polos_path):
                    shutil.copy2(pdf_polos_path, os.path.join(cert_folder, target_pdf_name))
        except Exception as e:
            print(f"Error copying bulk item to custom directory: {e}")

        # Save to DB
        cert_data_resmi['docx_path'] = docx_resmi_path
        if os.path.exists(pdf_resmi_path):
            cert_data_resmi['pdf_path'] = pdf_resmi_path
        db.save_certificate_db(cert_data_resmi)

        generated_records.append({
            'no_sertifikat': full_no,
            'nama': nama,
            'overall': cert_data_resmi['overall'],
            'qr_type': qr_type
        })

    zip_filename = f"PAKET_BULK_SERTIFIKAT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    zip_path = os.path.join(OUTPUT_DIR, zip_filename)

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(batch_dir):
            for file in files:
                abs_f = os.path.join(root, file)
                rel_f = os.path.relpath(abs_f, batch_dir)
                zipf.write(abs_f, arcname=rel_f)

    return {
        'success': True,
        'count': len(generated_records),
        'zip_filename': zip_filename,
        'zip_url': f"/api/v1/certificates/download/{zip_filename}",
        'gdrive_uploaded_count': gdrive_uploaded_count,
        'gdrive_skipped_count': gdrive_skipped_count,
        'gdrive_failed_count': gdrive_failed_count,
        'records': generated_records
    }
